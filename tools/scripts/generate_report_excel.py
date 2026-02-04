#!/usr/bin/env python3
"""
진단 결과 Excel 보고서 생성 스크립트

API 인벤토리(Task 2-1) 결과를 기준으로, 각 취약점 진단 결과(Task 2-2~2-5)를
엔드포인트별 Excel 보고서로 변환합니다.

사용법:
    python generate_report_excel.py <api_result> <finding_results...> --output <report.xlsx>
    python generate_report_excel.py state/pcona_task_21_result.json state/pcona_task_2*.json --output report.xlsx

입력:
    - Task 2-1 API 인벤토리 결과 (엔드포인트 목록)
    - Task 2-2~2-5 취약점 진단 결과 (findings)

출력:
    - 보고서 Excel 파일 (취약점_리스트, 인젝션, XSS, 파일처리, 데이터보호 탭)
"""

import json
import re
import sys
import argparse
from pathlib import Path
from datetime import date
from dataclasses import dataclass, field
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl 라이브러리가 필요합니다. pip install openpyxl")
    sys.exit(1)


# =============================================================================
#  상수 정의
# =============================================================================

# 점검 항목 카테고리 매핑
CATEGORY_MAP = {
    "injection": {
        "tab_name": "인젝션",
        "no_prefix": "1",
        "items": {
            "sql_injection": ("SQL인젝션", "DB정보 유출", "Risk 2"),
            "os_command": ("OS Command 인젝션", "서버 침투", "Risk 1"),
            "ssi_injection": ("SSI/SSTI 인젝션", "서버 침투", "Risk 1"),
            "nosql_injection": ("NoSQL 인젝션", "DB정보 유출", "Risk 2"),
        },
        "default_item": ("SQL인젝션", "DB정보 유출", "Risk 2"),
    },
    "xss": {
        "tab_name": "XSS",
        "no_prefix": "4",
        "items": {
            "persistent": ("Cross-Site Scripting: Persistent", "피싱/악성코드 배포", "Risk 3"),
            "reflected": ("Cross-Site Scripting: Reflected", "피싱/악성코드 배포", "Risk 3"),
            "dom": ("Cross-Site Scripting: DOM-based", "피싱/악성코드 배포", "Risk 3"),
            "redirect": ("Open Redirect", "피싱 유도", "Risk 3"),
        },
        "default_item": ("Cross-Site Scripting: Persistent", "피싱/악성코드 배포", "Risk 3"),
    },
    "file_handling": {
        "tab_name": "파일처리",
        "no_prefix": "5",
        "items": {
            "upload": ("파일업로드", "웹쉘 업로드 및 서버 내부 침투", "Risk 2"),
            "download": ("파일다운로드", "정보노출", "Risk 2"),
            "lfi": ("로컬 파일 인클루전", "서버 파일 노출", "Risk 2"),
            "path_traversal": ("경로 탐색", "서버 파일 노출", "Risk 2"),
        },
        "default_item": ("파일업로드", "웹쉘 업로드 및 서버 내부 침투", "Risk 2"),
    },
    "data_protection": {
        "tab_name": "데이터 보호",
        "no_prefix": "7",
        "items": {
            "info_leak": ("정보 누출", "정보 노출", "Risk 3"),
            "hardcoded_secret": ("하드코딩된 비밀정보", "계정 탈취", "Risk 2"),
            "cors": ("CORS 설정 미흡", "정보 노출", "Risk 3"),
            "jwt": ("JWT 취약점", "세션 탈취", "Risk 2"),
            "csrf": ("CSRF 보호 미흡", "위조 요청", "Risk 3"),
        },
        "default_item": ("정보 누출", "정보 노출", "Risk 3"),
    },
}

# Task ID → 카테고리 매핑
TASK_CATEGORY_MAP = {
    "2-2": "injection",
    "22": "injection",
    "task_22": "injection",
    "2-3": "xss",
    "23": "xss",
    "task_23": "xss",
    "2-4": "file_handling",
    "24": "file_handling",
    "task_24": "file_handling",
    "2-5": "data_protection",
    "25": "data_protection",
    "task_25": "data_protection",
}

# 점검 결과 매핑 (severity → 한글)
RESULT_MAP = {
    "critical": "취약",
    "high": "취약",
    "medium": "정보",
    "low": "양호",
    "info": "정보",
    "safe": "양호",
}

# Excel 스타일
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
VULN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
INFO_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SAFE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
#  데이터 클래스
# =============================================================================

@dataclass
class Endpoint:
    """API 엔드포인트"""
    method: str
    api: str
    handler: str = ""
    file: str = ""
    module: str = ""
    auth_required: bool = False
    parameters: list = field(default_factory=list)


@dataclass
class Finding:
    """취약점 발견 항목"""
    title: str
    severity: str
    category: str
    description: str = ""
    location: str = ""
    file: str = ""
    line: int = 0
    endpoint: str = ""
    recommendation: str = ""


@dataclass
class EndpointResult:
    """엔드포인트별 진단 결과"""
    endpoint: Endpoint
    result: str  # 취약, 양호, 정보
    item_name: str
    threat: str
    severity: str
    findings: list = field(default_factory=list)
    view_file: str = "N/A"
    process_file: str = ""


# =============================================================================
#  데이터 로딩
# =============================================================================

def load_api_inventory(filepath: Path) -> list[Endpoint]:
    """Task 2-1 API 인벤토리 로드"""
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)

    endpoints = []
    # scan_api.py 출력 형식
    if "endpoints" in data:
        for ep in data["endpoints"]:
            endpoints.append(Endpoint(
                method=ep.get("method", "GET"),
                api=ep.get("api", ""),
                handler=ep.get("handler", ""),
                file=ep.get("file", ""),
                module=ep.get("module", ""),
                auth_required=ep.get("auth_required", False),
                parameters=ep.get("parameters", []),
            ))
    # task_output_schema 형식
    elif "findings" in data:
        for f in data["findings"]:
            endpoints.append(Endpoint(
                method=f.get("method", "GET"),
                api=f.get("api", ""),
                handler=f.get("handler", ""),
                file=f.get("file", "").split(":")[0] if f.get("file") else "",
                module=f.get("module", ""),
                auth_required=f.get("auth_required", False),
                parameters=f.get("parameters", []),
            ))

    return endpoints


def load_findings(filepath: Path) -> tuple[str, list[Finding]]:
    """취약점 진단 결과 로드, 카테고리와 findings 반환"""
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)

    # task_id에서 카테고리 추출
    task_id = data.get("task_id", "")
    category = None
    for tid, cat in TASK_CATEGORY_MAP.items():
        if tid in task_id.lower() or tid in filepath.name.lower():
            category = cat
            break

    if not category:
        # 파일명에서 추출 시도
        fname = filepath.name.lower()
        if "injection" in fname or "22" in fname:
            category = "injection"
        elif "xss" in fname or "23" in fname:
            category = "xss"
        elif "file" in fname or "24" in fname:
            category = "file_handling"
        elif "data" in fname or "25" in fname:
            category = "data_protection"
        else:
            category = "injection"  # 기본값

    findings = []
    for f in data.get("findings", []):
        # 엔드포인트 추출 시도
        endpoint = ""
        location = f.get("location", {})
        if isinstance(location, dict):
            endpoint = location.get("endpoint", location.get("api", ""))
        elif isinstance(location, str):
            # 문자열에서 API 경로 추출
            match = re.search(r'(/\S+)', location)
            if match:
                endpoint = match.group(1)

        # affected_files에서 추출
        if not endpoint:
            for af in f.get("affected_files", []):
                if isinstance(af, dict) and "api" in af:
                    endpoint = af["api"]
                    break

        findings.append(Finding(
            title=f.get("title", ""),
            severity=f.get("severity", "info").lower(),
            category=f.get("category", ""),
            description=f.get("description", ""),
            location=str(location) if location else "",
            file=f.get("file", location.get("file", "") if isinstance(location, dict) else ""),
            line=f.get("line", location.get("line", 0) if isinstance(location, dict) else 0),
            endpoint=endpoint,
            recommendation=f.get("recommendation", ""),
        ))

    return category, findings


def match_findings_to_endpoints(
    endpoints: list[Endpoint],
    category: str,
    findings: list[Finding]
) -> list[EndpointResult]:
    """엔드포인트에 findings를 매칭하여 결과 생성"""

    cat_config = CATEGORY_MAP[category]
    results = []

    # 엔드포인트별로 결과 생성
    for idx, ep in enumerate(endpoints, 1):
        # 이 엔드포인트와 관련된 findings 찾기
        matched_findings = []
        for f in findings:
            # 엔드포인트 매칭
            if f.endpoint and f.endpoint in ep.api:
                matched_findings.append(f)
            # 파일 매칭
            elif f.file and ep.file and f.file in ep.file:
                matched_findings.append(f)
            # 핸들러 매칭
            elif ep.handler:
                handler_parts = ep.handler.replace("()", "").split(".")
                if any(p in f.title or p in f.location for p in handler_parts):
                    matched_findings.append(f)

        # 결과 판정
        if matched_findings:
            # 가장 높은 severity 기준
            severities = [RESULT_MAP.get(f.severity, "정보") for f in matched_findings]
            if "취약" in severities:
                result = "취약"
            elif "정보" in severities:
                result = "정보"
            else:
                result = "양호"

            # 점검 항목 결정 (첫 번째 finding 기준)
            first_finding = matched_findings[0]
            item_key = None
            for key, (name, _, _) in cat_config["items"].items():
                if key in first_finding.category.lower() or key in first_finding.title.lower():
                    item_key = key
                    break
            if item_key:
                item_name, threat, severity = cat_config["items"][item_key]
            else:
                item_name, threat, severity = cat_config["default_item"]

            view_file = first_finding.file if first_finding.file else "N/A"
        else:
            result = "양호"
            item_name, threat, severity = cat_config["default_item"]
            view_file = "N/A"

        results.append(EndpointResult(
            endpoint=ep,
            result=result,
            item_name=item_name,
            threat=threat,
            severity=severity,
            findings=matched_findings,
            view_file=view_file,
            process_file=ep.file,
        ))

    return results


# =============================================================================
#  Excel 생성
# =============================================================================

def create_detail_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    service_name: str,
    results: list[EndpointResult],
    no_prefix: str,
    check_date: str,
):
    """상세 탭 시트 생성 (인젝션, XSS, 파일처리, 데이터보호)"""

    ws = wb.create_sheet(sheet_name)

    # Row 1: 서비스명
    ws.merge_cells('B1:F1')
    ws['B1'] = f"[서비스명: {service_name}]"
    ws['B1'].font = Font(bold=True, size=12)

    # Row 2: 헤더
    headers = [
        "", "NO", "플랫폼구분", "점검항목", "점검결과", "점검일자",
        "조치계획", "조치내역", "조치일자", "이행점검결과",
        "발생 위협", "심각도", "View File", "RequestMapping", "Process File"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # 데이터 행
    row_num = 3
    for idx, r in enumerate(results, 1):
        no = f"{no_prefix}-{idx}"

        # 조치 관련 필드
        if r.result == "취약":
            action_plan = "ex) 조치완료/조치예정"
            action_detail = "ex) 상세 조치내용 기재"
            action_date = "ex) 21.00.00"
            verify_result = "-"
            result_fill = VULN_FILL
        elif r.result == "정보":
            action_plan = "N/A"
            action_detail = "N/A"
            action_date = "N/A"
            verify_result = "N/A"
            result_fill = INFO_FILL
        else:
            action_plan = "N/A"
            action_detail = "N/A"
            action_date = "N/A"
            verify_result = "N/A"
            result_fill = SAFE_FILL

        row_data = [
            "",  # A열 빈칸
            no,
            "WEB",
            r.item_name,
            r.result,
            check_date,
            action_plan,
            action_detail,
            action_date,
            verify_result,
            r.threat,
            r.severity,
            r.view_file[:50] if r.view_file else "N/A",
            r.endpoint.api,
            r.process_file[:50] if r.process_file else "N/A",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')
            # 점검결과 열 색상
            if col == 5:
                cell.fill = result_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_num += 1

    # 열 너비 조정
    col_widths = [3, 8, 10, 35, 10, 12, 20, 30, 12, 12, 20, 10, 30, 40, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_summary_sheet(
    wb: openpyxl.Workbook,
    service_name: str,
    all_results: dict[str, list[EndpointResult]],
    check_date: str,
):
    """취약점_리스트 요약 시트 생성"""

    ws = wb.create_sheet("취약점_리스트", 0)

    # Row 1: 서비스명
    ws.merge_cells('B1:F1')
    ws['B1'] = f"[서비스명: {service_name}]"
    ws['B1'].font = Font(bold=True, size=12)

    # Row 2: 설명
    ws.merge_cells('B2:L2')
    ws['B2'] = "진단 결과(인젝션, XSS, 파일처리, 데이터 보호) - 취약/정보 항목만 표시"

    # Row 3: 헤더
    headers = [
        "", "NO", "Sub_No", "플랫폼구분", "점검탭", "점검항목", "점검결과",
        "조치계획", "조치내역", "조치일자", "조치 Commit 번호", "이행점검결과"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # 취약/정보 항목만 수집
    row_num = 4
    total_no = 1

    for category, results in all_results.items():
        cat_config = CATEGORY_MAP[category]
        tab_name = cat_config["tab_name"]
        no_prefix = cat_config["no_prefix"]

        for idx, r in enumerate(results, 1):
            if r.result not in ("취약", "정보"):
                continue

            sub_no = f"{no_prefix}-{idx}"

            if r.result == "취약":
                action_plan = "ex) 조치완료/조치예정"
                action_detail = "ex) 상세 조치내용 기재"
                action_date = "ex) 21.00.00"
                action_commit = "ex) 00000000"
                verify_result = "-"
            else:
                action_plan = "N/A"
                action_detail = "N/A"
                action_date = "N/A"
                action_commit = "N/A"
                verify_result = "N/A"

            row_data = [
                "",
                total_no,
                sub_no,
                "WEB",
                tab_name,
                r.item_name,
                r.result,
                action_plan,
                action_detail,
                action_date,
                action_commit,
                verify_result,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical='center')
                if col == 7:  # 점검결과
                    if r.result == "취약":
                        cell.fill = VULN_FILL
                    else:
                        cell.fill = INFO_FILL
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            row_num += 1
            total_no += 1

    # 열 너비 조정
    col_widths = [3, 6, 10, 10, 12, 30, 10, 20, 30, 12, 18, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_url_sheet(wb: openpyxl.Workbook, endpoints: list[Endpoint]):
    """점검 URL 시트 생성 (API 인벤토리)"""

    ws = wb.create_sheet("점검 URL")

    # Row 1: 안내
    ws['A1'] = "API 엔드포인트 목록"
    ws['A1'].font = Font(bold=True)

    # Row 2: 헤더
    headers = ["Number", "Module", "Method", "URL", "Handler", "Auth", "Parameters"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # 데이터
    for idx, ep in enumerate(endpoints, 1):
        params = ", ".join(
            p.get("name", "") for p in ep.parameters
            if p.get("type") not in ("request", "response", "exchange", "session")
        )
        auth = "AUTH" if ep.auth_required else "OPEN"

        row_data = [idx, ep.module, ep.method, ep.api, ep.handler, auth, params]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 2, column=col, value=value)
            cell.border = THIN_BORDER

    # 열 너비
    col_widths = [8, 15, 8, 45, 40, 8, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_report(
    api_file: Path,
    finding_files: list[Path],
    output_file: Path,
    service_name: str = "서비스명",
):
    """최종 보고서 생성"""

    # API 인벤토리 로드
    endpoints = load_api_inventory(api_file)
    print(f"API 인벤토리: {len(endpoints)}개 엔드포인트")

    # Findings 로드 및 카테고리별 정리
    category_findings: dict[str, list[Finding]] = {}
    for fpath in finding_files:
        category, findings = load_findings(fpath)
        if category not in category_findings:
            category_findings[category] = []
        category_findings[category].extend(findings)
        print(f"  {fpath.name}: {len(findings)}건 ({category})")

    # 엔드포인트별 결과 매칭
    all_results: dict[str, list[EndpointResult]] = {}
    for category in CATEGORY_MAP.keys():
        findings = category_findings.get(category, [])
        results = match_findings_to_endpoints(endpoints, category, findings)
        all_results[category] = results

        # 통계
        vuln_count = sum(1 for r in results if r.result == "취약")
        info_count = sum(1 for r in results if r.result == "정보")
        safe_count = sum(1 for r in results if r.result == "양호")
        print(f"  {CATEGORY_MAP[category]['tab_name']}: 취약={vuln_count}, 정보={info_count}, 양호={safe_count}")

    # Excel 생성
    wb = openpyxl.Workbook()
    # 기본 시트 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    check_date = date.today().strftime("%y.%m.%d")

    # 요약 시트
    create_summary_sheet(wb, service_name, all_results, check_date)

    # 상세 시트들
    for category, results in all_results.items():
        cat_config = CATEGORY_MAP[category]
        create_detail_sheet(
            wb,
            cat_config["tab_name"],
            service_name,
            results,
            cat_config["no_prefix"],
            check_date,
        )

    # 점검 URL 시트
    create_url_sheet(wb, endpoints)

    # 저장
    wb.save(output_file)
    print(f"\n보고서 저장: {output_file}")


# =============================================================================
#  메인
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="진단 결과 Excel 보고서 생성"
    )
    parser.add_argument(
        "api_result",
        help="Task 2-1 API 인벤토리 결과 JSON 파일",
    )
    parser.add_argument(
        "finding_results",
        nargs="+",
        help="Task 2-2~2-5 취약점 진단 결과 JSON 파일들",
    )
    parser.add_argument(
        "--output", "-o",
        help="출력 Excel 파일 경로",
        default="진단결과_보고서.xlsx",
    )
    parser.add_argument(
        "--service", "-s",
        help="서비스명",
        default="서비스명",
    )
    args = parser.parse_args()

    api_file = Path(args.api_result)
    if not api_file.exists():
        print(f"Error: API 인벤토리 파일을 찾을 수 없습니다: {api_file}")
        sys.exit(1)

    finding_files = []
    for fpath in args.finding_results:
        p = Path(fpath)
        if p.exists():
            finding_files.append(p)
        else:
            print(f"Warning: 파일을 찾을 수 없습니다: {fpath}")

    if not finding_files:
        print("Error: 취약점 진단 결과 파일이 없습니다.")
        sys.exit(1)

    generate_report(
        api_file,
        finding_files,
        Path(args.output),
        args.service,
    )


if __name__ == "__main__":
    main()
